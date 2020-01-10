import pandas as pd
from datetime import timedelta, date
from openpyxl import load_workbook
import xlsxwriter
from copy import copy


def find_data(channel, sheet_name):
    global i, real_channel
    for i in range(len(channel)):
        if channel['Total Individuals'].values[i] < 0.95:
            real_channel = channel[:i]
            break
    real_channel.to_excel(excel_writer, sheet_name)


def copy_and_paste(sheet_number, sheet_name):
    global write_excel_file, i
    copy_data_file = load_workbook(r'Q:\홍희수\코바코\주간광고보고서\보고서\raw_data.xlsx')
    worksheet_copy = copy_data_file.worksheets[sheet_number]
    last_row = worksheet_copy.max_row
    row = 2
    col = 2
    array = []
    inner_array = []
    for i in range(last_row-1):  # copy
        for j in range(14):
            value = worksheet_copy.cell(row=row, column=col).value
            inner_array.append(value)
            col += 1

        array.append(inner_array)
        inner_array = []
        col = 2
        row += 1

    worksheet_write = write_excel_file[sheet_name]
    row = 7
    col = 2
    for i in range(last_row - 1):  # paste
        for j in range(14):
            test = worksheet_write.cell(row=row, column=col)
            test.value = array[i][j]
            if row >= 8:
                worksheet_write.cell(row=row, column=col)._style \
                    = copy(worksheet_write.cell(row=row-1, column=col)._style)

            col += 1
        col = 2
        row += 1
    row = 7
    col = 1
    for i in range(last_row-1):  # Numbering & Copying the style of the sheet
        worksheet_write.cell(row=row, column=col).value = i+1
        if row >= 8:
            worksheet_write.cell(row=row, column=col)._style \
                = copy(worksheet_write.cell(row=row - 1, column=col)._style)
        row += 1
        col = 1


# raw_data preprocessing(using data when it's upper 0.95)
excel_writer = pd.ExcelWriter(r'Q:\홍희수\코바코\주간광고보고서\보고서\raw_data.xlsx', engine='xlsxwriter')
kbs = pd.read_excel(r'Q:\홍희수\코바코\주간광고보고서\보고서\Raw_data.xls', sheet_name='수도권 전체 KBS2 GRP per Spot', skiprows=3)
mbc = pd.read_excel(r'Q:\홍희수\코바코\주간광고보고서\보고서\Raw_data.xls', sheet_name='수도권 전체 MBC GRP per Spot', skiprows=3)
sbs = pd.read_excel(r'Q:\홍희수\코바코\주간광고보고서\보고서\Raw_data.xls', sheet_name='수도권 전체 SBS GRP per Spot', skiprows=3)

# SBS
sbs_news_l = sbs[sbs['Programme'].values == 'SBS8뉴스L']
sbs_news_l = sbs_news_l.dropna(axis=1)
sbs_news = sbs[sbs['Programme'].values == 'SBS8뉴스']
sbs_news = sbs_news.dropna(axis=1)
weekdays = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
sbs_news['Day Of Week\Target'] = pd.Categorical(sbs_news['Day Of Week\Target'], categories=weekdays)
sbs_news = sbs_news.sort_values('Day Of Week\Target')
sbs_news_complete = pd.merge(sbs_news, sbs_news_l, on='Day Of Week\Target')
sbs_news_complete = sbs_news_complete.drop('Programme_y', axis=1)
sbs_news_complete = sbs_news_complete.rename(columns={'Programme_x': 'Programme'})
sbs = pd.concat([sbs, sbs_news_complete]).drop_duplicates(['Programme', 'Day Of Week\Target'], keep='last')
sbs = sbs.dropna(axis=0).sort_values('Total Individuals', ascending=False)

find_data(kbs, 'kbs')
find_data(mbc, 'mbc')
find_data(sbs, 'sbs')
excel_writer.save()
excel_writer.close()

# Copy and Paste to Excel File
write_excel_file = load_workbook(filename=r'Q:\홍희수\코바코\주간광고보고서\보고서\닐슨_코바코광고보고서.xlsx')
copy_and_paste(0, 'KBS2')
copy_and_paste(1, 'MBC')
copy_and_paste(2, 'SBS')

today = date.today()
last_monday = today - timedelta(days=today.weekday(), weeks=2)
last_sunday = (today - timedelta(days=today.weekday()) + timedelta(days=6, weeks=-1))
write_excel_file['KBS2'].cell(row=2, column=5).value = '%s ~ %s (PR, TJ, YG 기준)' % (last_monday, last_sunday)
write_excel_file.save(r'Q:\홍희수\코바코\주간광고보고서\보고서\닐슨_코바코광고보고서(%s-%s).xlsx'
                      % (last_monday.strftime('%y%m%d'), last_sunday.strftime('%y%m%d')))
