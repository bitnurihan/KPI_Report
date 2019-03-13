from datetime import datetime, date, timedelta
from openpyxl import load_workbook
import xlrd


def read_data_from_excel(write_file_name):
    global read_excel_file, worksheet_read, month, row, col, array, inner_array, i, j, value
    read_excel_file = xlrd.open_workbook(write_file_name)
    worksheet_read = read_excel_file.sheet_by_name('전체 수도권')
    month = datetime.now().month
    print(month)
    row = 5 + (month - 2)
    col = 3
    array = []
    inner_array = []
    for i in range(3):
        for j in range(36):
            value = worksheet_read.cell_value(rowx=row, colx=col)
            print(value)
            inner_array.append(value)
            col += 1

        array.append(inner_array)
        inner_array = []
        col = 3
        row += 1
        print()


def paste_to_excel(row, col, init_number):
    global i, j, test, test
    month = datetime.now().month
    print(month)
    for i in range(2):
        for j in range(36):
            test = worksheet_write.cell(row=row, column=col)
            test.value = array[i][j]
            col += 1

        if month < 4:
            month = 3
        elif month < 7:
            month = 4
        elif month < 11:
            month = 6
        else:
            month = 7

        row = init_number + (month - 2)
        col = 2
        print()

def prev_bounds(when=None):
    if not when : when = datetime.today()
    this_first = date(when.year, when.month, 1)
    prev_end = this_first - timedelta(days=1)
    prev_first = date(prev_end.year, prev_end.month, 1)
    return prev_first, prev_end

###1. 방송국 전시간대 시청률(평일 06:00-11:00, 17:00-24:00/ 주말 06:00-25:00)
write_excel_file = load_workbook(filename =r'C:\Users\hanbi01\Desktop\한빛누리\(매월)SBS월간업데이트\MonthlyReport1.xlsx')

read_data_from_excel(r'C:\Users\hanbi01\Desktop\한빛누리\(매월)SBS월간업데이트\1_3.xls')
worksheet_write = write_excel_file[r'2019년']

paste_to_excel(11 + (month - 2), 2, 28)


###2. 방송국 전시간대 시청률(06:00-25:00)
read_data_from_excel(r'C:\Users\hanbi01\Desktop\한빛누리\(매월)SBS월간업데이트\1_4.xls')
paste_to_excel(44 + (month - 2), 2, 61)


###3.SBS 프라임 시간대 시청률(평일 19:00-24:00, 주말 18:00-24:00)

read_data_from_excel(r'C:\Users\hanbi01\Desktop\한빛누리\(매월)SBS월간업데이트\1_5.xls')
paste_to_excel(77 + (month - 2), 2, 94)


###날짜 넣기

date_cell = worksheet_write['A3']
date_cell.value = "분석기간 : %s ~ %s" % (prev_bounds())
print(date_cell.value)


write_excel_file.save('testfile.xlsx')