import pyexcel
from openpyxl import load_workbook
import os
from copy import copy
from datetime import datetime


def copy_data_value_from_raw_data(write_file_name):
    global copy_data_file, copy_data_sheet, summary_row_list, row
    copy_data_file = load_workbook(write_file_name)
    copy_data_sheet = copy_data_file['(P) - sbs_kpi_']
    summary_row_list = []
    date_list = []
    total_duration = 0
    for row in range(1, 2500):
        channel_name = copy_data_sheet.cell(row=row + 4, column=1).value
        program_name = copy_data_sheet.cell(row=row + 4, column=2).value
        duration_sum = copy_data_sheet.cell(row=row + 4, column=15).value
        date_count = copy_data_sheet.cell(row=row + 4, column=5).value
        date_list.append(date_count)

        if duration_sum is None:  # sum duration cells
            pass
        else:
            total_duration += duration_sum

        if program_name is None:
            pass
        elif program_name.startswith('Summary'):  # find row the cell value starts with 'Summary' and copy program name
            dict_cntr = list(set(date_list))
            counting = len(dict_cntr) - dict_cntr.count(None)
            summary_row_list.append(dict(
                program_name=program_name[8:],
                indi_amr=copy_data_sheet.cell(row=row + 4, column=8).value,
                indi_shr=copy_data_sheet.cell(row=row + 4, column=9).value,
                young_amr=copy_data_sheet.cell(row=row + 4, column=10).value,
                young_shr=copy_data_sheet.cell(row=row + 4, column=11).value,
                hou_amr=copy_data_sheet.cell(row=row + 4, column=12).value,
                hou_shr=copy_data_sheet.cell(row=row + 4, column=13).value,
                duplication_count=counting,
                total_duration=total_duration - duration_sum)
            )
            total_duration = 0
            date_list = []

        if channel_name is None:
            pass
        elif channel_name.startswith('Summary'):  # find row the cell value starts with 'Summary' and copy program name
            dict_cntr = list(set(date_list))
            counting = len(dict_cntr) - dict_cntr.count(None)
            summary_row_list.append(dict(
                channel_name=channel_name[8:],
                indi_amr=copy_data_sheet.cell(row=row + 4, column=8).value,
                indi_shr=copy_data_sheet.cell(row=row + 4, column=9).value,
                young_amr=copy_data_sheet.cell(row=row + 4, column=10).value,
                young_shr=copy_data_sheet.cell(row=row + 4, column=11).value,
                hou_amr=copy_data_sheet.cell(row=row + 4, column=12).value,
                hou_shr=copy_data_sheet.cell(row=row + 4, column=13).value,
                duplication_count=counting,
                total_duration=total_duration - duration_sum)
            )
            total_duration = 0
            date_list = []


def paste_data_from_value():
    global row
    for dictionary in summary_row_list:
        channel = list(dictionary.values())

        for row in range(17, 100):
            title = paste_file_sheet.cell(row=row, column=2).value

            if title == channel[0]:
                col = 3
                for i in range(8):
                    paste_file_sheet.cell(row=row, column=col + i).value = channel[i + 1]


def paste_data_from_value_different_program_name(title_name):
    global row
    for dictionary in summary_row_list:
        channel = list(dictionary.values())

        for row in range(17, 100):
            title = paste_file_sheet.cell(row=row, column=2).value

            if title == title_name:
                col = 3
                for i in range(8):
                    paste_file_sheet.cell(row=row, column=col + i).value = channel[i + 1]


def paste_data_from_value_sbs_news():
    global row
    for dictionary in summary_row_list:
        channel = list(dictionary.values())

        for row in range(17, 80):
            title = paste_file_sheet.cell(row=row, column=2).value

            if title == 'SBS 8뉴스(평일)':
                if channel[0] == 'Work week':
                    col = 3
                    for i in range(8):
                        paste_file_sheet.cell(row=row, column=col + i).value = channel[i + 1]

            elif title == 'SBS 8뉴스(주말)':
                if channel[0] == 'Week end':
                    col = 3
                    for i in range(8):
                        paste_file_sheet.cell(row=row, column=col + i).value = channel[i + 1]


def copy_paste_annual_data(read_file_name):
    global copy_data_file, copy_data_sheet, col, array, inner_array, i, j, value, test
    copy_data_file = load_workbook(read_file_name)

    if datetime.now().month == 1:
        year = datetime.now().year - 1
    else:
        year = datetime.now().year

    copy_data_sheet = copy_data_file['(P) - 채널별시청률_']
    copy_data_sheet.delete_rows(20)
    copy_data_sheet.delete_rows(16)
    copy_data_sheet.delete_rows(13)
    
    # copy data from raw files
    row = 13
    col = 4
    array = []
    inner_array = []
    for i in range(6):
        for j in range(2):
            value = copy_data_sheet.cell(row=row, column=col).value
            inner_array.append(value)
            col += 1

        array.append(inner_array)
        inner_array = []
        col = 4
        row += 1
    
    # paste data to complete file
    row = 4
    col = 2

    for i in range(6):
        for j in range(2):
            test = paste_file_sheet.cell(row=row, column=col)
            test.value = array[i][j]
            col += 1
        row += 1
        if row == 6:
            row=8
        elif row == 7:
            row=9
        elif row == 10:
            row=12
        elif row == 11:
            row=13

        col = 2


def copy_paste_annual_data_primetime(read_file_name):
    global copy_data_file, copy_data_sheet, col, array, inner_array, i, j, value, test
    copy_data_file = load_workbook(read_file_name)
    copy_data_sheet = copy_data_file['(P) - 채널별시청률_']
    
    # deleted unusing rows
    copy_data_sheet.delete_rows(20)
    copy_data_sheet.delete_rows(16)
    copy_data_sheet.delete_rows(13)
    
    # copy data from raw files
    row = 13
    col = 4
    array = []
    inner_array = []
    for i in range(6):
        for j in range(2):
            value = copy_data_sheet.cell(row=row, column=col).value
            inner_array.append(value)
            col += 1

        array.append(inner_array)
        inner_array = []
        col = 4
        row += 1
    
    # paste data to complete file
    row = 4
    col = 4

    for i in range(6):
        for j in range(2):
            test = paste_file_sheet.cell(row=row, column=col)
            test.value = array[i][j]
            col += 1
        row += 1
        if row == 6:
            row=8
        elif row == 7:
            row=9
        elif row == 10:
            row=12
        elif row == 11:
            row=13

        col = 4


# *.xls -> *.xlsx
path_dir = r'C:\Users\hanbi01\Desktop\한빛누리\(분기)KPI\raw_data'
file_list = os.listdir(path_dir)
new_path_dir = r'C:\Users\hanbi01\Desktop\한빛누리\(분기)KPI\xlsx_raw_data'
new_file_list = os.listdir(new_path_dir)

for file in file_list:
    pyexcel.save_book_as(file_name=("%s\%s" % (path_dir, file)), dest_file_name="%s\%s.xlsx" % (new_path_dir, file.split('.')[0]))

# delete unusing cols
for file in new_file_list:
    new_file = load_workbook("%s\%s" % (new_path_dir, file))

    if os.path.splitext(file)[0] == '1_10':  # 1_10/ 1_11 파일은 해당 열을 지울 필요가 없음. pass 처리
        pass
    elif os.path.splitext(file)[0] == '1_11':
        pass
    else:
        sheet = new_file['(P) - sbs_kpi_']
        sheet.delete_cols(18)
        sheet.delete_cols(17)
        sheet.delete_cols(16)
        sheet.delete_cols(13)
        sheet.delete_cols(12)
        sheet.delete_cols(11)
        sheet.delete_cols(9)
        new_file.save("%s" % file)


# copy from data files
paste_file = load_workbook(r'C:\Users\hanbi01\Desktop\한빛누리\(분기)KPI\KPI.xlsx')
paste_file_sheet = paste_file['1']

copy_file_list = ['1.xlsx', '1_1.xlsx', '1_9.xlsx']

for copy_file in copy_file_list:
    copy_data_value_from_raw_data(copy_file)
    paste_data_from_value()

copy_program_list = [['1_2.xlsx', '정글의법칙'],
                     ['1_3.xlsx', 'SBS 월화드라마'],
                     ['1_4.xlsx', 'SBS 수목드라마'],
                     ['1_5.xlsx', '아침연속극'],
                     ['1_6.xlsx', 'SBS 금토드라마'],
                     ['1_8.xlsx', '모닝와이드 2부(평일)']]

for file_name, program_name in copy_program_list:
    copy_data_value_from_raw_data(file_name)
    paste_data_from_value_different_program_name(program_name)


copy_data_value_from_raw_data('1_7.xlsx')
paste_data_from_value_sbs_news()

copy_paste_annual_data(r'C:\Users\hanbi01\Desktop\한빛누리\(분기)KPI\xlsx_raw_data\1_10.xlsx')
copy_paste_annual_data_primetime(r'C:\Users\hanbi01\Desktop\한빛누리\(분기)KPI\xlsx_raw_data\1_11.xlsx')

# copy style in excel files
paste_file_style = load_workbook(r'C:\Users\hanbi01\Desktop\한빛누리\(분기)KPI\KPI.xlsx')
paste_file_sheet_style = paste_file['1']

row = 1
col = 1

for i in range(100):
    for j in range(12):
        paste_file_sheet.cell(row=row, column=col)._style = copy(
            paste_file_sheet_style.cell(row=row, column=col)._style)
        col += 1

    row += 1
    col = 1

# Saving complete file
if datetime.now().month == 1:
    year = datetime.now().year - 1
else:
    year = datetime.now().year
quarter = (datetime.now().month - 1) / 3

paste_file.save(r'C:\Users\hanbi01\Desktop\한빛누리\(분기)KPI\KPI_SBS %d년 %d분기_실적.xlsx' % (year, quarter))
