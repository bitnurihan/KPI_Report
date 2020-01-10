import pyexcel
from openpyxl import load_workbook
import os


def copy_data(file, sheet_number, range_i, range_j, start_row, start_col):
    copy_data_file = load_workbook(r'C:\Users\hanbi01\Desktop\한빛누리\Skylife_월간 보고서\raw_file\%s' % file)
    copy_data_sheet = copy_data_file.worksheets[sheet_number]
    last_row = copy_data_sheet.max_row
    last_col = copy_data_sheet.max_column
    row = start_row
    col = start_col
    if range_i == 100:
        last_row = 0
    elif range_j == 0:
        range_j = last_col
    array = []
    inner_array = []
    for i in range(last_row + range_i):
        for j in range(range_j):
            value = copy_data_sheet.cell(row=row, column=col).value
            inner_array.append(value)
            col += 1
        array.append(inner_array)
        inner_array = []
        col = start_col
        row += 1
    return array, last_col


def copy_and_paste_1st_sheet():
    global array
    array, last_col = copy_data('1.xlsx', 1, -3, 2, 4, 1)
    paste_data_sheet = paste_data_file.worksheets[0]
    row = 5
    col = 9
    for i in range(len(array)):
        cell_value = paste_data_sheet.cell(row=row, column=col)
        cell_value.value = array[i][0]
        row += 1

    col = 12
    row = 5
    for i in range(len(array)):
        cell_value = paste_data_sheet.cell(row=row, column=col)
        cell_value.value = array[i][1]
        row += 1


def copy_and_paste_rest_sheet(file, sheet_number):
    array, last_col = copy_data(file, 1, -1, 2, 4, 1)

    paste_data_sheet = paste_data_file.worksheets[sheet_number]
    row = 5
    col = 7
    for i in range(len(array)):
        for j in range(2):
            cell_value = paste_data_sheet.cell(row=row, column=col)
            cell_value.value = array[i][j]
            col += 1
        col = 7
        row += 1


def copy_and_paste_report_file(file, sheet_name, start_copy_row, start_paste_row):
    global last_col
    array, last_col = copy_data(file, 1, -4, 0, start_copy_row, 1)

    paste_data_sheet = paste_data_file[sheet_name]
    row = start_paste_row
    col = 2
    col_separate = int(len(array)/2)
    for i in range(len(array)):
        if i <= col_separate:
            for j in range(last_col):
                cell_value = paste_data_sheet.cell(row=row, column=col)
                cell_value.value = array[i][j]
                col += 1
            col = 2
            row += 1
        else:
            col = 8
            for j in range(last_col):
                cell_value = paste_data_sheet.cell(row=row-col_separate-1, column=col)
                cell_value.value = array[i][j]
                col += 1
            row += 1


def copy_and_paste_to_main_file(file, sheet_number, sheet_name, start_row, start_column):
    array, last_col = copy_data(file, sheet_number, -3, 0, 4, 1)

    paste_data_sheet = paste_data_file[sheet_name]
    row = start_row
    col = start_column
    for i in range(len(array)):
        for j in range(last_col):
            cell_value = paste_data_sheet.cell(row=row, column=col)
            cell_value.value = array[i][j]
            col += 1
        col = start_column
        row += 1


def program_sheet_copy_last_month(sheet_number, sheet_name):
    array, last_col = copy_data('닐슨_SkyLife분기보고서(2019년4분기)_누리.xlsx', sheet_number, 100, 3, 8, 1)
    paste_data_sheet = paste_data_file[sheet_name]
    row = 8
    col = 5
    for i in range(100):
        for j in range(3):
            cell_value = paste_data_sheet.cell(row=row, column=col)
            cell_value.value = array[i][j]
            col += 1
        col = 5
        row += 1


def program_sheet_paste_this_month(file, sheet_name):
    array, last_col = copy_data(file, 1, 100, 3, 4, 1)

    paste_data_sheet = paste_data_file[sheet_name]
    row = 8
    col = 1
    for i in range(100):
        for j in range(3):
            cell_value = paste_data_sheet.cell(row=row, column=col)
            cell_value.value = array[i][j]
            col += 1
        col = 1
        row += 1


def data_input_to_main_file(file, sheet_name, sheet_number, paste_column, start_row, add_counter, range_i):
    global i
    for i in range(range_i):
        copy_and_paste_to_main_file(file, sheet_number, sheet_name, start_row, paste_column)
        sheet_number += 1
        paste_column += add_counter


# *.xls -> *.xlsx
path_dir = r'C:\Users\hanbi01\Desktop\한빛누리\Skylife_월간 보고서\data'
file_list = os.listdir(path_dir)
new_path_dir = r'C:\Users\hanbi01\Desktop\한빛누리\Skylife_월간 보고서\raw_file'
new_file_list = os.listdir(new_path_dir)
main_file_name = r'C:\Users\hanbi01\Desktop\한빛누리\Skylife_월간 보고서\닐슨_SkyLife분기보고서(2019년4분기)_누리.xlsx'
main_preparing_file_name = r'C:\Users\hanbi01\Desktop\한빛누리\Skylife_월간 보고서\SkyLife분기작업용.xlsx'

for file in file_list:
    pyexcel.save_book_as(file_name=("%s\%s" % (path_dir, file)), dest_file_name="%s\%s.xlsx" % (new_path_dir, file.split('.')[0]))

paste_data_file = load_workbook(main_preparing_file_name)
copy_and_paste_1st_sheet()  # Skylife

file_list = ['1_13.xlsx', '1_3.xlsx', '1_11.xlsx']  # [Cable, IPTV, Audio Channel]
number = 1
for file in file_list:
    copy_and_paste_rest_sheet(file, number)
    number += 1
paste_data_file.save(main_preparing_file_name)

paste_data_file = load_workbook(main_file_name)

copy_and_paste_report_file('1_1.xlsx', '2-2.채널별시청동향(상세분석)', 5, 8)
copy_and_paste_report_file('1_12.xlsx', '8-2.케이블채널별시청동향-2', 4, 7)
copy_and_paste_report_file('1_4.xlsx', '9-2.IPTV채널별시청동향-2', 4, 7)

data_input_to_main_file('1_6.xlsx', '4.개인남자여자', 1, 2, 9, 3, 3)
data_input_to_main_file('1_7.xlsx', '4.성연령별', 1, 2, 8, 2, 5)
data_input_to_main_file('1_7.xlsx', '4.성연령별', 6, 2, 226, 2, 5)
data_input_to_main_file('1_8.xlsx', '5.소득별', 1, 2, 8, 2, 7)
data_input_to_main_file('1_9.xlsx', '5.직업별', 1, 2, 8, 2, 8)
data_input_to_main_file('1_10.xlsx', '5.지역별', 1, 2, 8, 2, 5)

program_sheet_copy_last_month(10, '8-3.케이블프로그램')
program_sheet_copy_last_month(13, '9-3.IPTV프로그램')

program_sheet_paste_this_month('1_2.xlsx', '8-3.케이블프로그램')
program_sheet_paste_this_month('1_5.xlsx', '9-3.IPTV프로그램')

paste_data_file.save(main_file_name)
