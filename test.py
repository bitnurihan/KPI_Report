from openpyxl import load_workbook
from copy import copy
from datetime import datetime


def copy_data_value_from_raw_data(write_file_name):
    global copy_data_file, copy_data_sheet, summary_row_list, row
    copy_data_file = load_workbook(write_file_name)
    copy_data_sheet = copy_data_file['전체 수도권']
    summary_row_list = []
    date_list = []
    total_duration = 0
    for row in range(1, 501):
        channel_name = copy_data_sheet.cell(row=row + 4, column=1).value
        program_name = copy_data_sheet.cell(row=row + 4, column=2).value
        duration_sum = copy_data_sheet.cell(row=row + 4, column=15).value
        date_count = copy_data_sheet.cell(row=row + 4, column=5).value
        date_list.append(date_count)

        if duration_sum is None:  # duration 셀 더하기
            pass
        else:
            total_duration += duration_sum

        if program_name is None:
            pass
        elif program_name.startswith('Summary'):  # Summary로 시작하는 행 찾아서 프로그램명 넣고 row 찾기
            dict_cntr = list(set(date_list))
            counting = len(dict_cntr) - dict_cntr.count(None)
            summary_row_list.append(dict(
                program_name=program_name[8:],
                indi_amr=copy_data_sheet.cell(row=row + 4, column=8).value,
                indi_shr=copy_data_sheet.cell(row=row + 4, column=9).value,
                young_amr=copy_data_sheet.cell(row=row + 4, column=10).value,
                young_shr=copy_data_sheet.cell(row=row + 4, column=11).value,
                hou_amr=copy_data_sheet.cell(row=row + 4, column=12).value,
                hou_shr=copy_data_sheet.cell(row=row + 4, column=13).value,
                duplication_count=counting,
                total_duration=total_duration - duration_sum)
            )
            total_duration = 0
            date_list = []

        if channel_name is None:
            pass
        elif channel_name.startswith('Summary'):  # Summary로 시작하는 행 찾아서 프로그램명 넣고 row 찾기
            dict_cntr = list(set(date_list))
            counting = len(dict_cntr) - dict_cntr.count(None)
            summary_row_list.append(dict(
                channel_name=channel_name[8:],
                indi_amr=copy_data_sheet.cell(row=row + 4, column=8).value,
                indi_shr=copy_data_sheet.cell(row=row + 4, column=9).value,
                young_amr=copy_data_sheet.cell(row=row + 4, column=10).value,
                young_shr=copy_data_sheet.cell(row=row + 4, column=11).value,
                hou_amr=copy_data_sheet.cell(row=row + 4, column=12).value,
                hou_shr=copy_data_sheet.cell(row=row + 4, column=13).value,
                duplication_count=counting,
                total_duration=total_duration - duration_sum)
            )
            total_duration = 0
            date_list = []


def paste_data_from_value():
    global row
    for dictionary in summary_row_list:
        channel = list(dictionary.values())

        for row in range(17, 100):
            title = paste_file_sheet.cell(row=row, column=2).value

            if title == channel[0]:
                col = 3
                for i in range(8):
                    paste_file_sheet.cell(row=row, column=col + i).value = channel[i + 1]


def paste_data_from_value_different_program_name(title_name):
    global row
    for dictionary in summary_row_list:
        channel = list(dictionary.values())
        print()

        for row in range(17, 100):
            title = paste_file_sheet.cell(row=row, column=2).value

            if title == title_name:
                col = 3
                for i in range(8):
                    paste_file_sheet.cell(row=row, column=col + i).value = channel[i + 1]

                if title_name == '주말드라마':  # 주말드라마는 하루 4회 방영으로 *2 해주어야 함
                    counting = paste_file_sheet.cell(row=row, column=9)
                    counting.value = int(counting.value) * 2


def paste_data_from_value_sbs_news():
    global row



paste_file = load_workbook(r'C:\Users\hanbi01\Desktop\한빛누리\(분기)KPI\KPI.xlsx')
paste_file_sheet = paste_file['4분기']
# paste_file_sheet = paste_file['%d분기' % ((datetime.now().month - 1) / 3 + 1)]

copy_data_value_from_raw_data('1.xlsx')

paste_data_from_value()

copy_data_value_from_raw_data('1_1.xlsx')

paste_data_from_value()

copy_data_value_from_raw_data('1_9.xlsx')

paste_data_from_value()

copy_data_value_from_raw_data('1_2.xlsx')

paste_data_from_value_different_program_name('정글의법칙')

copy_data_value_from_raw_data('1_3.xlsx')

paste_data_from_value_different_program_name('SBS 월화드라마')

copy_data_value_from_raw_data('1_4.xlsx')

paste_data_from_value_different_program_name('SBS 수목드라마')

copy_data_value_from_raw_data('1_5.xlsx')

paste_data_from_value_different_program_name('아침연속극')

copy_data_value_from_raw_data('1_6.xlsx')

paste_data_from_value_different_program_name('주말드라마')

copy_data_value_from_raw_data('1_7.xlsx')

for dictionary in summary_row_list:
    channel = list(dictionary.values())

    for row in range(17, 80):
        title = paste_file_sheet.cell(row=row, column=2).value

        if title == 'SBS 8뉴스(평일)':
            if channel[0] == 'Work week':
                col = 3
                for i in range(8):
                    paste_file_sheet.cell(row=row, column=col + i).value = channel[i + 1]

        elif title == 'SBS 8뉴스(주말)':
            if channel[0] == 'Week end':
                col = 3
                for i in range(8):
                    paste_file_sheet.cell(row=row, column=col + i).value = channel[i + 1]



copy_data_value_from_raw_data('1_8.xlsx')
paste_data_from_value_different_program_name('모닝와이드 2부(평일)')

# 망가진 스타일 복사
paste_file_style = load_workbook(r'C:\Users\hanbi01\Desktop\한빛누리\(분기)KPI\KPI.xlsx')
paste_file_sheet_style = paste_file['4분기']
# paste_file_sheet_style = paste_file_style['%d분기' % ((datetime.now().month - 1) / 3 + 1)]

row = 1
col = 1

for i in range(100):
    for j in range(12):
        paste_file_sheet.cell(row=row, column=col)._style = copy(
            paste_file_sheet_style.cell(row=row, column=col)._style)
        col += 1

    row += 1
    col = 1

paste_file.save('testfile3.xlsx')
